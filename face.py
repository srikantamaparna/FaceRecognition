import cv2
import openpyxl
from simple_facerec import SimpleFacerec
sfr = SimpleFacerec()
sfr.load_encoding_images("C:/Users/srika/Desktop/new/")

cap = cv2.VideoCapture(0)

while True:
    n=1
    while n==1:
        ret, frame = cap.read()

        face_locations, face_names = sfr.detect_known_faces(frame)
        for face_loc, name in zip(face_locations, face_names):
            y1, x2, y2, x1 = face_loc[0], face_loc[1], face_loc[2], face_loc[3]
            cv2.putText(frame, name, (x1, y1 - 10), cv2.FONT_HERSHEY_DUPLEX, 1, (0, 0, 200), 2)
            cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 0, 200), 4)
            if name=='Unknown':
                print("Unkown")
                exit(0)
            if len(name)!=0:
                c = name[8::]
                n=0
    wrkbk = openpyxl.load_workbook("C:/Users/srika/Desktop/excel2.xlsx")
    sh = wrkbk.active
    for j in range(1, sh.max_column + 1):
        cell_obj = sh.cell(row=int(c), column=j)
        if j!=sh.max_column :
            print(cell_obj.value, end="-")
        else:
            print(cell_obj.value)

    break
    cv2.imshow("Frame", frame)
    key = cv2.waitKey(1)
    if key & 0xFF == ord('q'):
        break
cap.release()
cv2.destroyAllWindows()
